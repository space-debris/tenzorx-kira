"""
Paytm Statement Generator for Kirana Loan Demo
Generates realistic fake Paytm UPI statements — both PDF and XLSX.

Usage:
    python generate_paytm_statement.py

Outputs saved to OUTPUT_DIR (default: current directory):
    Paytm_Statement_MAR26.pdf
    Paytm_Statement_MAR26.xlsx
"""

import os
import random
import string
from datetime import datetime, timedelta

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ─────────────────────────────────────────────────────────────
# CONFIG — edit these to customise the generated profile
# ─────────────────────────────────────────────────────────────

CONFIG = {
    "owner_name":       "RAMESH KUMAR",
    "phone":            "9876543210",
    "email":            "ramesh.kirana@gmail.com",
    "bank":             "State Bank Of India",
    "account_suffix":   "47",
    "month_label":      "MAR'26",
    "month_start":      datetime(2026, 3, 1),
    "month_end":        datetime(2026, 3, 31),
    "num_transactions": 80,
    "incoming_ratio":   0.99,    # 99% incoming, 1% outgoing
    "profile":          "medium",  # "small" | "medium" | "large"
    "output_dir":       ".",
}

# Kirana revenue profiles: typical per-transaction incoming amount (INR)
PROFILES = {
    "small":  {"min": 20,   "max": 600},
    "medium": {"min": 80,   "max": 3000},
    "large":  {"min": 300,  "max": 12000},
}

CUSTOMER_FIRST = ["Suresh","Priya","Amit","Neha","Ravi","Sunita","Mohit","Kavita",
                  "Deepak","Pooja","Arjun","Meera","Vikas","Anjali","Rohit","Sushma",
                  "Naveen","Rakesh","Dinesh","Geeta","Lalit","Kiran","Sunil","Seema",
                  "Manish","Usha","Satish","Poonam","Tarun","Nisha"]
CUSTOMER_LAST  = ["Sharma","Verma","Singh","Gupta","Kumar","Yadav","Joshi","Mishra",
                  "Rani","Nair","Pillai","Tiwari","Sinha","Dubey","Pandey","Reddy",
                  "Agarwal","Jain","Kaur","Patel","Mehta","Rao","Iyer","Bansal",
                  "Malhotra","Bala","Chaudhary","Saxena","Thakur","Chauhan"]

OUTGOING_VENDORS = [
    ("HUL Distributor",   "huldist{r}@okaxis"),
    ("ITC Distributor",   "itcdist{r}@ybl"),
    ("Amul Dairy",        "amulsupply{r}@okicici"),
    ("Airtel Recharge",   "airtel.recharge@ptyes"),
    ("BSES Electricity",  "bses{r}@okhdfc"),
    ("Nestle Dist.",      "nestledist{r}@okaxis"),
]

INCOMING_TAGS = ["", "", "", "#Shopping", "#Daily Needs", "#Food", "#Grocery"]
OUTGOING_TAGS = ["#Supplier", "#Recharge", "#Utilities"]


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def rand_ref():
    return ''.join(random.choices(string.digits, k=12))

def rand_order():
    return ''.join(random.choices(string.digits, k=11))

def rand_upi(name, r):
    slug = name.lower().replace(" ", "")[:8]
    domains = ["@oksbi", "@okaxis", "@ybl", "@okicici", "@paytm", "@okhdfcbank"]
    return f"{slug}{r}{random.choice(domains)}"

def random_dt(start, end):
    delta = end - start
    dt = start + timedelta(seconds=random.randint(0, int(delta.total_seconds())))
    return dt.replace(hour=random.randint(7, 21),
                      minute=random.randint(0, 59),
                      second=random.randint(0, 59))

def gen_transactions(cfg):
    profile = PROFILES[cfg["profile"]]
    txns = []
    for _ in range(cfg["num_transactions"]):
        incoming = random.random() < cfg["incoming_ratio"]
        dt = random_dt(cfg["month_start"], cfg["month_end"])
        r  = random.randint(10, 99)

        if incoming:
            name = f"{random.choice(CUSTOMER_FIRST)} {random.choice(CUSTOMER_LAST)}"
            upi  = rand_upi(name, r)
            desc = f"Received from {name}"
            other = f"{upi} on Paytm"
            amt  = round(random.uniform(profile["min"], profile["max"]))
            tag  = random.choice(INCOMING_TAGS)
        else:
            vname, upi_tmpl = random.choice(OUTGOING_VENDORS)
            upi  = upi_tmpl.format(r=r)
            desc = f"To {vname}"
            other = f"{upi} on Paytm"
            amt  = -round(random.uniform(500, 8000))
            tag  = random.choice(OUTGOING_TAGS)

        txns.append({
            "dt":    dt,
            "desc":  desc,
            "other": other,
            "acct":  f"{cfg['bank']} - {cfg['account_suffix']}",
            "amt":   amt,
            "ref":   rand_ref(),
            "order": rand_order(),
            "tag":   tag,
        })

    txns.sort(key=lambda x: x["dt"])
    return txns

def compute_summary(txns):
    paid  = [t for t in txns if t["amt"] < 0]
    recv  = [t for t in txns if t["amt"] >= 0]
    return {
        "total_paid":  abs(sum(t["amt"] for t in paid)),
        "paid_count":  len(paid),
        "total_recv":  sum(t["amt"] for t in recv),
        "recv_count":  len(recv),
    }


# ─────────────────────────────────────────────────────────────
# XLSX — exactly matches Paytm format (2 sheets)
# ─────────────────────────────────────────────────────────────

def make_xlsx(txns, s, cfg):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    for col, w in zip("ABCDE", [75, 28, 26, 28, 26]):
        ws1.column_dimensions[col].width = w

    def cell(r, c, val, bold=False, size=10, color="000000", bg=None, align="left", wrap=False):
        cl = ws1.cell(row=r, column=c, value=val)
        cl.font = Font(bold=bold, size=size, color=color)
        cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            cl.fill = PatternFill("solid", fgColor=bg)
        return cl

    # Blank rows
    for r in range(1, 5):
        ws1.row_dimensions[r].height = 6

    cell(5, 1, cfg["owner_name"],   bold=True, size=13, color="0070BA")
    ws1.row_dimensions[5].height = 18
    cell(6, 1, cfg["phone"])
    cell(7, 1, cfg["email"])

    ws1.row_dimensions[8].height = 8

    period = f"1 {cfg['month_label']} - 31 {cfg['month_label']}"
    cell(9, 1, "Paytm Statement for :", bold=True)
    cell(9, 2, period)
    ws1.row_dimensions[9].height = 16

    cell(10, 1, "Money Paid (Amount in Rs.)", bold=True)
    cell(10, 2, f"-{s['total_paid']:.2f}")
    cell(11, 1, "Money Paid (No. of Payments)", bold=True)
    cell(11, 2, s["paid_count"])
    cell(12, 1, "Money Received (Amount in Rs.)", bold=True)
    cell(12, 2, f"+{s['total_recv']:.2f}")
    cell(13, 1, "Money Received (No. of Payments)", bold=True)
    cell(13, 2, s["recv_count"])
    for r in range(10, 14):
        ws1.row_dimensions[r].height = 16

    ws1.row_dimensions[14].height = 8

    hdrs14 = ["Accounts", "Money Paid\n(Amount in Rs.)", "Money Paid\n(No. of Payments)",
              "Money Received\n(Amount in Rs.)", "Money Received\n(No. of Payments)"]
    for i, h in enumerate(hdrs14, 1):
        cell(15, i, h, bold=True, bg="D9E8F5", align="center", wrap=True)
    ws1.row_dimensions[15].height = 30

    cell(16, 1, f"{cfg['bank']} - {cfg['account_suffix']}", align="center")
    acct_vals = [s["total_paid"], s["paid_count"], s["total_recv"], s["recv_count"]]
    for i, v in enumerate(acct_vals, 2):
        cl = ws1.cell(row=16, column=i, value=v)
        cl.alignment = Alignment(horizontal="center")
        if i in (2, 4):
            cl.number_format = "#,##0.00"
    ws1.row_dimensions[16].height = 18

    ws1.row_dimensions[17].height = 8
    cell(18, 1, "All payments done by you on Paytm App are reflected in this statement")
    cell(19, 1, "Notes:", bold=True)
    notes = [
        "- Self transfer payments are not included in the total money paid and money received calculations.",
        "- Payments that you might have hidden on payment history page will not be included in this statement.",
        "- Payments that you might have done using Paytm Payments Bank Wallet will not be included in this statement",
    ]
    for i, n in enumerate(notes, 20):
        cell(i, 1, n)
        ws1.row_dimensions[i].height = 16

    ws1.row_dimensions[23].height = 8
    cell(24, 1, "In case of queries,")
    cl = ws1.cell(row=24, column=2, value="contact us")
    cl.font = Font(color="0070BA", underline="single")

    # ── Sheet 2: Passbook Payment History ──
    ws2 = wb.create_sheet("Passbook Payment History")
    for col, w in zip("ABCDEFGHIJK", [14, 12, 35, 42, 28, 12, 16, 16, 12, 16, 12]):
        ws2.column_dimensions[col].width = w

    ph_hdrs = ["Date", "Time", "Transaction Details",
               "Other Transaction Details (UPI ID or A/c No)",
               "Your Account", "Amount", "UPI Ref No.", "Order ID",
               "Remarks", "Tags", "Comment"]
    for i, h in enumerate(ph_hdrs, 1):
        cl = ws2.cell(row=1, column=i, value=h)
        cl.font = Font(bold=True, size=9)
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cl.fill = PatternFill("solid", fgColor="D9E8F5")
    ws2.row_dimensions[1].height = 22

    for ri, t in enumerate(txns, 2):
        vals = [t["dt"].strftime("%d/%m/%Y"), t["dt"].strftime("%H:%M:%S"),
                t["desc"], t["other"], t["acct"], t["amt"],
                t["ref"], t["order"], "", t["tag"], ""]
        for ci, v in enumerate(vals, 1):
            cl = ws2.cell(row=ri, column=ci, value=v)
            cl.alignment = Alignment(horizontal="left", vertical="center")
        amt_cl = ws2.cell(row=ri, column=6)
        amt_cl.number_format = "#,##0.00"
        amt_cl.font = Font(color="1A7A1A" if t["amt"] >= 0 else "CC0000")
        ws2.row_dimensions[ri].height = 15

    tag = cfg["month_label"].replace("'", "")
    out = os.path.join(cfg["output_dir"], f"Paytm_Statement_{tag}.xlsx")
    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────
# PDF — Paytm-style visual layout
# ─────────────────────────────────────────────────────────────

BLUE  = colors.HexColor("#0070BA")
LIGHT = colors.HexColor("#E8F4FB")
GREEN = colors.HexColor("#1A7A1A")
RED   = colors.HexColor("#CC0000")
GREY  = colors.HexColor("#CCCCCC")
DGREY = colors.HexColor("#555555")

def make_pdf(txns, s, cfg):
    tag   = cfg["month_label"].replace("'", "")
    fpath = os.path.join(cfg["output_dir"], f"Paytm_Statement_{tag}.pdf")

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=14*mm)

    st = {
        "title":  ParagraphStyle("t1", fontName="Helvetica-Bold",  fontSize=14, textColor=BLUE),
        "sub":    ParagraphStyle("t2", fontName="Helvetica",        fontSize=8.5,textColor=DGREY),
        "label":  ParagraphStyle("t3", fontName="Helvetica-Bold",  fontSize=8.5,textColor=colors.black),
        "note":   ParagraphStyle("t4", fontName="Helvetica",        fontSize=7.5,textColor=DGREY, spaceAfter=1),
        "sec":    ParagraphStyle("t5", fontName="Helvetica-Bold",  fontSize=9,  textColor=colors.black, spaceAfter=2),
        "thdr":   ParagraphStyle("t6", fontName="Helvetica-Bold",  fontSize=7,  textColor=colors.white, alignment=TA_CENTER),
        "tcell":  ParagraphStyle("t7", fontName="Helvetica",        fontSize=6.5,textColor=colors.black, leading=9),
        "tin":    ParagraphStyle("t8", fontName="Helvetica-Bold",  fontSize=7,  textColor=GREEN, alignment=TA_RIGHT),
        "tout":   ParagraphStyle("t9", fontName="Helvetica-Bold",  fontSize=7,  textColor=RED,   alignment=TA_RIGHT),
    }

    story = []
    period = f"1 {cfg['month_label']} - 31 {cfg['month_label']}"

    # Header
    hdr = Table([[
        Paragraph("<b><font color='#0070BA'>paytm</font> <font color='#E91E8C'>♥</font> <font color='#0070BA'>UPI</font></b>", st["title"]),
        Paragraph(f"<b>{cfg['owner_name']}</b><br/>{cfg['phone']}, {cfg['email']}", st["sub"]),
    ]], colWidths=[90*mm, 90*mm])
    hdr.setStyle(TableStyle([
        ("ALIGN",(1,0),(1,0),"RIGHT"), ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    story += [hdr, HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=4)]

    # Summary boxes
    def pluralize(n, word="Payment"):
        return f"{n} {word}{'s' if n != 1 else ''}"

    sumbox = Table([[
        Paragraph(f"<b>Paytm Statement for</b><br/><font size=10><b>{period}</b></font>", st["label"]),
        Paragraph(f"<b>Total Money Paid</b><br/><font size=12 color='#CC0000'><b>- Rs.{s['total_paid']:,.0f}</b></font><br/><font size=8>{pluralize(s['paid_count'])} made</font>", st["label"]),
        Paragraph(f"<b>Total Money Received</b><br/><font size=12 color='#1A7A1A'><b>+ Rs.{s['total_recv']:,.0f}</b></font><br/><font size=8>{pluralize(s['recv_count'])} received</font>", st["label"]),
    ]], colWidths=[60*mm, 62*mm, 62*mm])
    sumbox.setStyle(TableStyle([
        ("BOX",(0,0),(0,0),0.5,GREY), ("BOX",(1,0),(1,0),0.5,GREY), ("BOX",(2,0),(2,0),0.5,GREY),
        ("BACKGROUND",(0,0),(0,0),LIGHT),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),8),
    ]))
    story += [sumbox, Spacer(1, 4*mm)]

    # Notes
    story.append(Paragraph("<b>Note:</b>", st["note"]))
    for n in ["Self transfer payments are not included in total money paid and received calculations",
              "Payments hidden on payment history page will not be included in this statement",
              "Payments done using Paytm Payments Bank Wallet will not be included in this statement"]:
        story.append(Paragraph(f"• {n}", st["note"]))
    story.append(Spacer(1, 3*mm))

    # Accounts table
    atbl = Table([[
        Paragraph("Accounts", st["thdr"]),
        Paragraph("Payment made", st["thdr"]),
        Paragraph("Payment received", st["thdr"]),
    ],[
        Paragraph(f"{cfg['bank']} - {cfg['account_suffix']}", st["tcell"]),
        Paragraph(f"Rs.{s['total_paid']:,.0f}\n({pluralize(s['paid_count'])})", st["tcell"]),
        Paragraph(f"Rs.{s['total_recv']:,.0f}\n({pluralize(s['recv_count'])})", st["tcell"]),
    ]], colWidths=[70*mm, 57*mm, 57*mm])
    atbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("BOX",(0,0),(-1,-1),0.5,GREY),
        ("INNERGRID",(0,0),(-1,-1),0.3,GREY),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6), ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story += [atbl, Spacer(1, 5*mm)]

    # Passbook
    story.append(Paragraph("<b>Passbook Payments History</b>", st["sec"]))
    story.append(Paragraph("All payments done by you on Paytm App are reflected in this statement", st["note"]))
    story.append(Spacer(1, 2*mm))

    col_w = [22*mm, 30*mm, 22*mm, 24*mm, 18*mm, 14*mm, 20*mm, 14*mm]
    txn_data = [[Paragraph(h, st["thdr"]) for h in
                 ["Date & Time","Transaction Details","Notes & Tags",
                  "Your Account","Amount","UPI Ref","Order ID","Remarks"]]]

    for t in txns:
        amt = t["amt"]
        row = [
            Paragraph(f"{t['dt'].strftime('%d %b')}\n{t['dt'].strftime('%I:%M %p')}", st["tcell"]),
            Paragraph(t["desc"], st["tcell"]),
            Paragraph(t["tag"] or "-", st["tcell"]),
            Paragraph(t["acct"], st["tcell"]),
            Paragraph(f"{'+ Rs.' if amt>=0 else '- Rs.'}{abs(amt):,.0f}", st["tin"] if amt>=0 else st["tout"]),
            Paragraph(t["ref"], st["tcell"]),
            Paragraph(t["order"], st["tcell"]),
            Paragraph("-", st["tcell"]),
        ]
        txn_data.append(row)

    ttbl = Table(txn_data, colWidths=col_w, repeatRows=1)
    ttbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LIGHT]),
        ("BOX",(0,0),(-1,-1),0.4,GREY), ("INNERGRID",(0,0),(-1,-1),0.2,GREY),
        ("TOPPADDING",(0,0),(-1,-1),3), ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),3), ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story += [ttbl, Spacer(1, 5*mm)]

    # Footer
    story.append(HRFlowable(width="100%", thickness=0.5, color=GREY))
    story.append(Spacer(1, 2*mm))
    ftbl = Table([[
        Paragraph("Page 1 of 1  |  Powered by UPI", st["note"]),
        Paragraph("For any queries, <font color='#0070BA'><u>Contact Us</u></font>", st["note"]),
    ]], colWidths=[95*mm, 95*mm])
    ftbl.setStyle(TableStyle([("ALIGN",(1,0),(1,0),"RIGHT")]))
    story.append(ftbl)

    doc.build(story)
    return fpath


# ─────────────────────────────────────────────────────────────
# Run
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    txns   = gen_transactions(CONFIG)
    summ   = compute_summary(txns)

    xlsx_p = make_xlsx(txns, summ, CONFIG)
    pdf_p  = make_pdf(txns, summ, CONFIG)

    print(f"Generated {len(txns)} transactions  |  Profile: {CONFIG['profile'].upper()}")
    print(f"  XLSX -> {xlsx_p}")
    print(f"  PDF  -> {pdf_p}")
    print(f"  Total Received : Rs.{summ['total_recv']:,.0f}  ({summ['recv_count']} txns)")
    print(f"  Total Paid     : Rs.{summ['total_paid']:,.0f}  ({summ['paid_count']} txns)")
